from fastapi import FastAPI, APIRouter, HTTPException, Query, Request, Response, Depends, UploadFile, File, Form
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional, Dict, Any, Literal
import uuid
from datetime import datetime, timezone, date, timedelta
from dateutil.relativedelta import relativedelta
from enum import Enum
import bcrypt
import jwt
import io
import re
import csv
import math
import json
import hashlib
from dateutil import parser as date_parser

from xlsx_simple import read_first_sheet_as_dict_rows

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]
ENABLE_BULK_ACTUALS = os.environ.get("ENABLE_BULK_ACTUALS", "true").lower() in {"1", "true", "yes", "on"}

app = FastAPI()
api_router = APIRouter(prefix="/api")

def ensure_bulk_actuals_enabled():
    if not ENABLE_BULK_ACTUALS:
        raise HTTPException(status_code=404, detail="Bulk actual imports are disabled")

# ============== AUTH HELPERS ==============
JWT_SECRET = os.environ.get("JWT_SECRET", "changeme")
JWT_ALGORITHM = "HS256"

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")

def verify_password(plain: str, hashed: str) -> bool:
    return bcrypt.checkpw(plain.encode("utf-8"), hashed.encode("utf-8"))

def create_access_token(user_id: str, email: str) -> str:
    payload = {"sub": user_id, "email": email, "exp": datetime.now(timezone.utc) + timedelta(days=7), "type": "access"}
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        auth_header = request.headers.get("Authorization", "")
        if auth_header.startswith("Bearer "):
            token = auth_header[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access":
            raise HTTPException(status_code=401, detail="Invalid token type")
        user = await db.users.find_one({"id": payload["sub"]})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        return {"id": user["id"], "email": user["email"], "name": user.get("name", "")}
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

class LoginRequest(BaseModel):
    email: str
    password: str

# ============== ENUMS ==============
class Category(str, Enum):
    REVENUE = "Revenue"
    SALARY = "Salary"
    TAX = "Tax"
    DEBT = "Debt"
    EXPENSE = "Expense"
    TRANSFER = "Transfer"
    COGS = "COGS"
    OTHER = "Other"

class Certainty(str, Enum):
    MATERIALIZED = "Materialized"
    SURE = "Sure to happen"
    FIFTY_FIFTY = "50/50"
    IDEA = "Idea"

class Recurrence(str, Enum):
    NONE = "none"
    MONTHLY = "monthly"
    QUARTERLY = "quarterly"

class RecurrenceMode(str, Enum):
    REPEAT = "repeat"
    DISTRIBUTE = "distribute"

class SourceType(str, Enum):
    MANUAL = "manual"
    DEAL = "deal"

class FlowPriority(str, Enum):
    CRITICAL = "critical"
    FLEXIBLE = "flexible"
    STRATEGIC = "strategic"

# ============== ENTITY MODELS ==============
class Entity(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    description: str = ""
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class EntityCreate(BaseModel):
    name: str
    description: str = ""

class EntityUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None

# ============== BANK ACCOUNT MODELS ==============
class BankAccount(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    entity_id: str = ""
    entity: str = ""
    label: str
    amount: float
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    updated_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class BankAccountCreate(BaseModel):
    entity_id: str
    label: str
    amount: float

class BankAccountUpdate(BaseModel):
    entity_id: Optional[str] = None
    label: Optional[str] = None
    amount: Optional[float] = None

# ============== TREASURY DEBT VIEW MODELS ==============
class DebtConsolidationItem(BaseModel):
    creditor: str
    entity: str
    entity_id: str
    total_debt_chf: float
    monthly_payment_chf: float
    frequency: str
    calculation_basis: str
    source_flow_id: str

class TreasuryDebtUpdate(BaseModel):
    creditor: Optional[str] = None
    total_debt_chf: Optional[float] = None

# ============== CASH FLOW MODELS ==============
class CashFlow(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    label: str
    amount: float
    date: str  # ISO date YYYY-MM-DD
    category: Category = Category.EXPENSE
    certainty: Certainty = Certainty.MATERIALIZED
    recurrence: Recurrence = Recurrence.NONE
    recurrence_mode: RecurrenceMode = RecurrenceMode.REPEAT
    recurrence_end: Optional[str] = None
    recurrence_count: Optional[int] = None
    entity_id: str = ""
    entity: str = ""
    parent_id: Optional[str] = None
    is_percentage: bool = False
    percentage_of_parent: Optional[float] = None
    carryover_from: Optional[str] = None
    carryover_month: Optional[str] = None
    source_type: SourceType = SourceType.MANUAL
    source_id: Optional[str] = None
    priority: Optional[FlowPriority] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    updated_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class CashFlowCreate(BaseModel):
    label: str
    amount: float
    date: str
    category: Category = Category.EXPENSE
    certainty: Certainty = Certainty.MATERIALIZED
    recurrence: Recurrence = Recurrence.NONE
    recurrence_mode: RecurrenceMode = RecurrenceMode.REPEAT
    recurrence_end: Optional[str] = None
    recurrence_count: Optional[int] = None
    entity_id: str
    parent_id: Optional[str] = None
    is_percentage: bool = False
    percentage_of_parent: Optional[float] = None
    source_type: SourceType = SourceType.MANUAL
    source_id: Optional[str] = None
    priority: Optional[FlowPriority] = None

class CashFlowUpdate(BaseModel):
    label: Optional[str] = None
    amount: Optional[float] = None
    date: Optional[str] = None
    category: Optional[Category] = None
    certainty: Optional[Certainty] = None
    recurrence: Optional[Recurrence] = None
    recurrence_mode: Optional[RecurrenceMode] = None
    recurrence_end: Optional[str] = None
    recurrence_count: Optional[int] = None
    entity_id: Optional[str] = None
    priority: Optional[FlowPriority] = None

class CashFlowBatchCreate(BaseModel):
    parent: CashFlowCreate
    linked: List[CashFlowCreate] = []

# ============== FLOW OCCURRENCE MODELS (ACTUALS & VARIANCE) ==============
class FlowOccurrence(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    flow_id: str
    month: str  # "YYYY-MM"
    actual_amount: Optional[float] = None  # None = no actual recorded yet (planned)
    variance_action: Optional[str] = None  # "carry_forward" | "write_off" | None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class FlowOccurrenceUpdate(BaseModel):
    flow_id: str
    month: str
    actual_amount: Optional[float] = None
    variance_action: Optional[str] = None  # "carry_forward" | "write_off"

class ActualImportRow(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    batch_id: str
    row_index: int
    include: bool = True
    status: str = "ready"  # ready | warning | discarded | applied | failed | skipped
    raw_date: str = ""
    raw_description: str = ""
    raw_amount: str = ""
    transaction_date: str  # YYYY-MM-DD
    month: str  # YYYY-MM
    description: str
    amount: float
    category: Category
    entity_id: Optional[str] = None
    suggested_flow_id: Optional[str] = None
    selected_flow_id: Optional[str] = None
    match_score: float = 0.0
    match_reason: str = ""
    variance_action: str = "actual_only"  # actual_only | carry_forward | write_off
    error: Optional[str] = None
    # existing_flow: map to selected_flow_id. new_flow: apply creates a new cash flow line from this row, then records the actual.
    classification: str = "existing_flow"
    # Per row: override replaces actual for that flow/month; addition adds this row's amount onto current actual.
    actual_merge_mode: Literal["override", "addition"] = "override"
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    updated_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class ActualImportBatch(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    filename: str
    file_type: str
    status: str = "draft"  # draft | applied | discarded | partial
    entity_id: Optional[str] = None
    total_rows: int = 0
    ready_rows: int = 0
    warning_rows: int = 0
    discarded_rows: int = 0
    applied_rows: int = 0
    failed_rows: int = 0
    skipped_rows: int = 0
    last_apply_fingerprint: Optional[str] = None
    last_applied_at: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    updated_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class ActualImportRowUpdate(BaseModel):
    model_config = ConfigDict(extra="forbid")
    include: Optional[bool] = None
    transaction_date: Optional[str] = None
    month: Optional[str] = None
    description: Optional[str] = None
    amount: Optional[float] = None
    category: Optional[Category] = None
    entity_id: Optional[str] = None
    selected_flow_id: Optional[str] = None
    variance_action: Optional[str] = None
    classification: Optional[Literal["existing_flow", "new_flow"]] = None
    actual_merge_mode: Optional[Literal["override", "addition"]] = None

class ActualImportApplyRequest(BaseModel):
    model_config = ConfigDict(extra="forbid")
    idempotency_key: Optional[str] = None
    # Fallback when a row has no actual_merge_mode (legacy rows); prefer per-row on ActualImportRow.
    actual_merge_mode: Literal["override", "addition"] = "override"


# ============== UNDO SYSTEM ==============
class UndoAction(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    action_type: str  # "create" | "update" | "delete" | "batch_create" | "record_actual" | "record_actual_batch"
    collection: str  # "cash_flows" | "bank_accounts" | "entities" | "flow_occurrences"
    target_ids: List[str] = []  # IDs affected
    previous_data: Optional[List[dict]] = None  # Snapshots before change
    created_data: Optional[List[dict]] = None  # Data that was created (for undoing creates)
    side_effects: Optional[dict] = None  # Related changes (carryover flows, occurrences)
    timestamp: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    description: str = ""

# ============== SETTINGS & PROJECTION MODELS ==============
class Settings(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = "settings"
    safety_buffer: float = 50000.0

class SettingsUpdate(BaseModel):
    safety_buffer: float

class MonthProjection(BaseModel):
    month: str
    month_label: str
    inflows: float
    outflows: float
    net: float
    closing_cash: float
    status: str

class ProjectionResponse(BaseModel):
    cash_now: float
    lowest_cash: float
    lowest_cash_month: str
    highest_pressure_month: str = ""
    first_watch_month: Optional[str] = None
    first_danger_month: Optional[str] = None
    overall_status: str
    safety_buffer: float
    months: List[MonthProjection]

# ============== AMOUNT SIGN NORMALIZATION ==============
def normalize_amount_for_category(category: str, amount: float) -> float:
    """Canonical sign rule: Revenue is positive; all other categories are negative."""
    if isinstance(category, Enum):
        category = category.value
    if amount == 0:
        return 0.0
    return abs(amount) if category == Category.REVENUE.value else -abs(amount)

def parse_import_amount(raw_value: Any) -> Optional[float]:
    if raw_value is None:
        return None
    if isinstance(raw_value, (int, float)):
        if isinstance(raw_value, float) and math.isnan(raw_value):
            return None
        return float(raw_value)

    text = str(raw_value).strip()
    if not text:
        return None
    text = text.replace("CHF", "").replace(" ", "")
    text = text.replace("'", "").replace("\u2019", "")

    if "," in text and "." in text:
        # Keep decimal separator nearest to end.
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")

    text = re.sub(r"[^0-9.\-]", "", text)
    if text in {"", "-", ".", "-."}:
        return None
    try:
        return float(text)
    except ValueError:
        return None

def best_flow_match(
    flows: List[dict],
    description: str,
    amount: float,
    entity_id: Optional[str] = None
) -> Dict[str, Any]:
    def tokens(text: str) -> set:
        return {t for t in re.split(r"\W+", (text or "").lower()) if len(t) > 2}

    desc_lower = (description or "").lower().strip()
    desc_tokens = tokens(desc_lower)
    amount_abs = abs(amount)
    amount_sign = 1 if amount > 0 else -1
    best: Dict[str, Any] = {"flow_id": None, "score": 0.0, "reason": "No close match"}

    for flow in flows:
        if entity_id and flow.get("entity_id") != entity_id:
            continue
        score = 0.0
        reasons = []

        flow_amount = float(flow.get("amount", 0))
        flow_sign = 1 if flow_amount > 0 else -1
        if flow_sign == amount_sign:
            score += 0.4
            reasons.append("direction")

        flow_abs = abs(flow_amount)
        if amount_abs > 0:
            diff_ratio = abs(flow_abs - amount_abs) / amount_abs
            if diff_ratio <= 0.05:
                score += 0.35
                reasons.append("amount-close")
            elif diff_ratio <= 0.15:
                score += 0.2
                reasons.append("amount-near")

        flow_label = (flow.get("label", "") or "").lower().strip()
        flow_tokens = tokens(flow_label)
        if flow_tokens and desc_tokens:
            overlap = len(flow_tokens.intersection(desc_tokens))
            union = len(flow_tokens.union(desc_tokens))
            jaccard = (overlap / union) if union > 0 else 0
            if jaccard >= 0.6:
                score += 0.35
                reasons.append("label-strong")
            elif jaccard >= 0.35:
                score += 0.22
                reasons.append("label-medium")
            elif overlap > 0:
                score += 0.1
                reasons.append("label-weak")
        elif flow_label and flow_label in desc_lower:
            score += 0.2
            reasons.append("label-substring")

        if score > best["score"]:
            best = {
                "flow_id": flow.get("id"),
                "score": round(min(score, 0.99), 3),
                "reason": ", ".join(reasons) if reasons else "weak",
            }

    return best

def build_apply_fingerprint(
    rows: List[dict],
    idempotency_key: Optional[str] = None,
) -> str:
    items = []
    for row in rows:
        if not row.get("include", True):
            continue
        cls = row.get("classification") or "existing_flow"
        # new_flow: fingerprint must stay stable after apply stores selected_flow_id on the row.
        flow_fp = None if cls == "new_flow" else row.get("selected_flow_id")
        row_merge = row.get("actual_merge_mode") or "override"
        items.append({
            "row_id": row.get("id"),
            "flow_id": flow_fp,
            "month": row.get("month"),
            "amount": round(float(row.get("amount", 0)), 2),
            "variance_action": row.get("variance_action", "actual_only"),
            "classification": cls,
            "actual_merge_mode": row_merge,
        })
    payload = {
        "rows": sorted(items, key=lambda x: x["row_id"]),
        "idempotency_key": idempotency_key or "",
    }
    raw = json.dumps(payload, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def build_cash_flow_from_import_row(row: dict, entity_id: str, amount_value: float) -> CashFlow:
    """Create a non-recurring planned line from an import row (used when classification is new_flow)."""
    label = (row.get("description") or "Imported").strip()[:200] or "Imported"
    raw_cat = row.get("category", Category.EXPENSE.value)
    cat_str = raw_cat.value if isinstance(raw_cat, Category) else str(raw_cat)
    try:
        cat_enum = Category(cat_str)
    except ValueError:
        cat_enum = Category.EXPENSE

    td = str(row.get("transaction_date") or "").strip()
    month = str(row.get("month") or "").strip()
    start_date: str
    if len(td) >= 10 and td[4] == "-" and td[7] == "-":
        start_date = td[:10]
    elif is_valid_month_key(month):
        start_date = f"{month}-01"
    else:
        start_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    norm_amt = normalize_amount_for_category(cat_enum.value, float(amount_value))
    return CashFlow(
        label=label,
        amount=norm_amt,
        date=start_date,
        category=cat_enum,
        certainty=Certainty.MATERIALIZED,
        recurrence=Recurrence.NONE,
        recurrence_mode=RecurrenceMode.REPEAT,
        entity_id=entity_id,
    )


def detect_import_columns(columns: List[str]) -> Dict[str, str]:
    cols: Dict[str, str] = {}
    for c in columns:
        if c is None:
            continue
        raw = str(c).strip()
        if not raw:
            continue
        key = raw.lower().replace("\ufeff", "").replace("\xa0", " ")
        cols[key] = raw

    date_exact = [
        "date", "transaction date", "booking date", "value date", "buchungsdatum", "buchungstag",
        "valuta", "valutadatum", "transaktionsdatum", "posting date", "trade date", "datum",
    ]
    desc_exact = [
        "description", "label", "details", "memo", "narrative", "name", "text", "beschreibung",
        "buchungstext", "verwendungszweck", "zweck", "payee", "empfänger", "empfaenger",
        "begünstigter", "beguenstigter", "merchant", "note", "bemerkung",
        "posting text", "buchungstext en", "booking text",
    ]
    amount_exact = [
        "amount", "value", "chf", "betrag", "amount chf", "transaction amount", "booked amount",
    ]
    debit_exact = ["debit", "soll", "belastung", "lastschrift", "abgang", "withdrawal"]
    credit_exact = ["credit", "haben", "gutschrift", "guthaben", "eingang", "deposit"]

    def pick_exact(keys: List[str]) -> Optional[str]:
        for k in keys:
            if k in cols:
                return cols[k]
        return None

    def pick_contains(subs: tuple, exclude: tuple = ()) -> Optional[str]:
        for col_lower, orig in cols.items():
            if any(ex in col_lower for ex in exclude):
                continue
            for sub in subs:
                if sub in col_lower:
                    return orig
        return None

    detected: Dict[str, str] = {}
    detected["date"] = pick_exact(date_exact) or pick_contains(
        (
            "buchungsdatum", "buchungstag", "valutadatum", "valuta", "transaktionsdatum",
            "booking date", "value date", "trade date", "posting date",
        ),
        ("betrag", "amount", "belastung", "gutschrift", "text", "beschreibung", "posting text"),
    )
    detected["description"] = pick_exact(desc_exact) or pick_contains(
        ("beschreibung", "text", "zweck", "verwendung", "detail", "memo", "narrative", "payee", "merchant"), ()
    )
    detected["amount"] = pick_exact(amount_exact) or pick_contains(
        ("betrag", "amount", "chf", "saldo"), ("datum", "date", "buchung")
    )

    if "amount" not in detected or detected["amount"] is None:
        debit_col = pick_exact(debit_exact) or pick_contains(("belastung", "debit", "soll", "abgang", "lastschrift"), ())
        credit_col = pick_exact(credit_exact) or pick_contains(("gutschrift", "guthaben", "credit", "haben", "eingang"), ())
        if debit_col and credit_col:
            detected["debit"] = debit_col
            detected["credit"] = credit_col
            detected.pop("amount", None)

    detected = {k: v for k, v in detected.items() if v}
    return detected


def decode_csv_bytes(content: bytes) -> str:
    for enc in ("utf-8-sig", "utf-8", "cp1252", "iso-8859-1", "mac_roman"):
        try:
            return content.decode(enc)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="replace")


def sniff_csv_delimiter(sample: str) -> str:
    first = ""
    for line in sample.splitlines():
        if line.strip():
            first = line
            break
    if not first:
        return ","
    if first.count("\t") > first.count(",") and first.count("\t") > first.count(";"):
        return "\t"
    if first.count(";") > first.count(","):
        return ";"
    return ","


def parse_csv_to_rows(content: bytes) -> tuple[List[str], List[Dict[str, Any]]]:
    text = decode_csv_bytes(content)
    delim = sniff_csv_delimiter(text[:8192])
    reader = csv.DictReader(io.StringIO(text), delimiter=delim)
    fieldnames = list(reader.fieldnames or [])
    rows = list(reader)
    return fieldnames, rows


def parse_import_row_date(raw: Any) -> Optional[date]:
    if raw is None or raw == "":
        return None
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    if isinstance(raw, (int, float)) and not isinstance(raw, bool):
        serial = float(raw)
        if 200 < serial < 100000:
            try:
                return date(1899, 12, 30) + timedelta(days=int(serial))
            except (OverflowError, ValueError):
                pass
    try:
        return date_parser.parse(str(raw)).date()
    except Exception:
        return None

def is_valid_month_key(month: str) -> bool:
    if not isinstance(month, str) or not re.match(r"^\d{4}-\d{2}$", month):
        return False
    year = int(month[:4])
    mon = int(month[5:7])
    return 1 <= mon <= 12 and 1900 <= year <= 3000

def normalize_import_transaction_date(raw_date: Optional[str]) -> Optional[str]:
    if raw_date is None:
        return None
    text = str(raw_date).strip()
    if not text:
        return None
    try:
        parsed = date_parser.parse(text)
        return parsed.strftime("%Y-%m-%d")
    except Exception:
        return None

async def validate_selected_flow_for_row(selected_flow_id: Optional[str], row_entity_id: Optional[str], batch_entity_id: Optional[str]) -> Optional[dict]:
    if not selected_flow_id:
        return None
    flow = await db.cash_flows.find_one({"id": selected_flow_id}, {"_id": 0, "id": 1, "entity_id": 1})
    if not flow:
        raise HTTPException(status_code=400, detail="Selected flow does not exist")

    # Row-level entity wins so bulk imports can mix entities in one batch.
    enforced_entity_id = row_entity_id or batch_entity_id
    if enforced_entity_id and flow.get("entity_id") != enforced_entity_id:
        raise HTTPException(status_code=400, detail="Selected flow does not belong to the import entity scope")
    return flow

# ============== HELPER: Calculate percentage-based amount dynamically ==============
async def get_flow_with_dynamic_amount(flow: dict) -> dict:
    """For percentage-based linked flows, calculate amount from parent dynamically."""
    if flow.get("is_percentage") and flow.get("parent_id") and flow.get("percentage_of_parent"):
        parent = await db.cash_flows.find_one({"id": flow["parent_id"]}, {"_id": 0})
        if parent:
            pct = flow["percentage_of_parent"]
            parent_abs = abs(parent.get("amount", 0))
            computed = parent_abs * pct / 100
            flow["amount"] = normalize_amount_for_category(
                flow.get("category", Category.EXPENSE.value),
                computed
            )
    return flow


# ============== UNDO HELPERS ==============
MAX_UNDO_STACK = 50

async def push_undo(action_type: str, collection: str, target_ids: List[str], 
                     previous_data=None, created_data=None, description="", side_effects=None):
    """Push an action to the undo stack."""
    action = UndoAction(
        action_type=action_type,
        collection=collection,
        target_ids=target_ids,
        previous_data=previous_data,
        created_data=created_data,
        description=description,
        side_effects=side_effects,
    )
    await db.undo_stack.insert_one(action.model_dump())
    # Trim stack to max size
    count = await db.undo_stack.count_documents({})
    if count > MAX_UNDO_STACK:
        oldest = await db.undo_stack.find({}, {"_id": 0, "id": 1}).sort("timestamp", 1).limit(count - MAX_UNDO_STACK).to_list(100)
        if oldest:
            await db.undo_stack.delete_many({"id": {"$in": [o["id"] for o in oldest]}})

async def restore_record_actual_side_effects(side_effects: dict):
    prev_occurrence = side_effects.get("previous_occurrence")
    flow_id = side_effects.get("flow_id", "")
    month = side_effects.get("month", "")

    if prev_occurrence:
        prev_occurrence.pop("_id", None)
        await db.flow_occurrences.replace_one(
            {"flow_id": flow_id, "month": month},
            prev_occurrence,
            upsert=True,
        )
    else:
        await db.flow_occurrences.delete_one({"flow_id": flow_id, "month": month})

    created_carryovers = side_effects.get("created_carryover_ids", [])
    for cid in created_carryovers:
        await db.cash_flows.delete_one({"id": cid})

    deleted_carryovers = side_effects.get("deleted_carryovers", [])
    for dc in deleted_carryovers:
        dc.pop("_id", None)
        await db.cash_flows.insert_one(dc)

    created_flow_id = side_effects.get("created_flow_id")
    if created_flow_id:
        await db.cash_flows.delete_many({"parent_id": created_flow_id})
        await db.cash_flows.delete_one({"id": created_flow_id})

@api_router.post("/undo")
async def undo_last_action():
    """Undo the most recent action. Restores full system state including dependencies."""
    last = await db.undo_stack.find_one({}, {"_id": 0}, sort=[("timestamp", -1)])
    if not last:
        return {"status": "nothing_to_undo"}
    
    action_type = last["action_type"]
    collection = last["collection"]
    coll = db[collection]
    side_effects = last.get("side_effects") or {}
    
    if action_type in ["create", "batch_create"]:
        # Delete the created items
        for tid in last.get("target_ids", []):
            await coll.delete_one({"id": tid})
        # For batch_create, also delete linked children
        if action_type == "batch_create" and collection == "cash_flows":
            for tid in last.get("target_ids", []):
                await coll.delete_many({"parent_id": tid})
    
    elif action_type == "update":
        # Restore previous data
        for prev in (last.get("previous_data") or []):
            prev_clean = {k: v for k, v in prev.items() if k != "_id"}
            fid = prev_clean.get("id")
            if fid:
                await coll.replace_one({"id": fid}, prev_clean, upsert=True)
    
    elif action_type == "delete":
        # Re-insert previously deleted data
        for prev in (last.get("previous_data") or []):
            prev.pop("_id", None)
            await coll.insert_one(prev)
        # If children were orphaned (not deleted), restore their parent_id
        orphaned_ids = side_effects.get("orphaned_child_ids", [])
        if orphaned_ids and collection == "cash_flows":
            parent_data = (last.get("previous_data") or [{}])[0]
            parent_id = parent_data.get("id", "")
            if parent_id:
                for child_snapshot in (last.get("previous_data") or [])[1:]:
                    child_id = child_snapshot.get("id")
                    if child_id in orphaned_ids:
                        # Restore the child to its pre-delete state (including parent_id)
                        child_clean = {k: v for k, v in child_snapshot.items() if k != "_id"}
                        await coll.replace_one({"id": child_id}, child_clean, upsert=True)
    
    elif action_type == "record_actual":
        await restore_record_actual_side_effects(side_effects)

    elif action_type == "record_actual_batch":
        # Undo in reverse apply order to restore nested carryover dependencies correctly.
        for row_side_effects in reversed(side_effects.get("rows", [])):
            await restore_record_actual_side_effects(row_side_effects)
    
    # Remove the action from stack
    await db.undo_stack.delete_one({"id": last["id"]})
    
    return {"status": "undone", "description": last.get("description", "")}

@api_router.get("/undo/peek")
async def peek_undo():
    """Peek at the last undoable action."""
    last = await db.undo_stack.find_one({}, {"_id": 0}, sort=[("timestamp", -1)])
    if not last:
        return {"has_undo": False}
    return {"has_undo": True, "description": last.get("description", ""), "action_type": last["action_type"]}

# ============== ENTITY ROUTES ==============
@api_router.get("/entities", response_model=List[Entity])
async def get_entities():
    entities = await db.entities.find({}, {"_id": 0}).to_list(100)
    return entities

@api_router.post("/entities", response_model=Entity)
async def create_entity(entity: EntityCreate):
    entity_obj = Entity(**entity.model_dump())
    await db.entities.insert_one(entity_obj.model_dump())
    return entity_obj

@api_router.put("/entities/{entity_id}", response_model=Entity)
async def update_entity(entity_id: str, update: EntityUpdate):
    update_data = {k: v for k, v in update.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="No valid update data")
    result = await db.entities.find_one_and_update(
        {"id": entity_id}, {"$set": update_data},
        return_document=True, projection={"_id": 0}
    )
    if not result:
        raise HTTPException(status_code=404, detail="Entity not found")
    return result

@api_router.delete("/entities/{entity_id}")
async def delete_entity(entity_id: str):
    accounts = await db.bank_accounts.count_documents({"entity_id": entity_id})
    flows = await db.cash_flows.count_documents({"entity_id": entity_id})
    if accounts > 0 or flows > 0:
        raise HTTPException(status_code=400, detail=f"Cannot delete: {accounts} accounts, {flows} flows")
    result = await db.entities.delete_one({"id": entity_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Entity not found")
    return {"message": "Entity deleted"}

# ============== BANK ACCOUNT ROUTES ==============
@api_router.get("/bank-accounts", response_model=List[BankAccount])
async def get_bank_accounts(entity_id: Optional[str] = None):
    query = {"entity_id": entity_id} if entity_id else {}
    accounts = await db.bank_accounts.find(query, {"_id": 0}).to_list(100)
    return accounts

@api_router.post("/bank-accounts", response_model=BankAccount)
async def create_bank_account(account: BankAccountCreate):
    entity = await db.entities.find_one({"id": account.entity_id})
    if not entity:
        raise HTTPException(status_code=400, detail="Entity not found")
    account_obj = BankAccount(**account.model_dump())
    await db.bank_accounts.insert_one(account_obj.model_dump())
    return account_obj

@api_router.put("/bank-accounts/{account_id}", response_model=BankAccount)
async def update_bank_account(account_id: str, update: BankAccountUpdate):
    update_data = {k: v for k, v in update.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="No valid update data")
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    result = await db.bank_accounts.find_one_and_update(
        {"id": account_id}, {"$set": update_data},
        return_document=True, projection={"_id": 0}
    )
    if not result:
        raise HTTPException(status_code=404, detail="Bank account not found")
    return result

@api_router.delete("/bank-accounts/{account_id}")
async def delete_bank_account(account_id: str):
    result = await db.bank_accounts.delete_one({"id": account_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Bank account not found")
    return {"message": "Bank account deleted"}

@api_router.get("/treasury/debts", response_model=List[DebtConsolidationItem])
async def get_treasury_debts(entity_id: Optional[str] = None):
    """Derived debt consolidation view from existing debt-tagged cash flows."""
    query = {"category": Category.DEBT.value}
    if entity_id:
        query["entity_id"] = entity_id

    flows = await db.cash_flows.find(query, {"_id": 0}).to_list(5000)
    if not flows:
        return []

    entity_ids = list({f.get("entity_id") for f in flows if f.get("entity_id")})
    entity_map = {}
    if entity_ids:
        entities = await db.entities.find({"id": {"$in": entity_ids}}, {"_id": 0, "id": 1, "name": 1}).to_list(5000)
        entity_map = {e["id"]: e["name"] for e in entities}

    debts = []
    for flow in flows:
        recurrence = flow.get("recurrence", Recurrence.NONE.value)
        recurrence_mode = flow.get("recurrence_mode", RecurrenceMode.REPEAT.value)
        recurrence_count = flow.get("recurrence_count")
        amount = abs(float(flow.get("amount", 0)))

        if recurrence == Recurrence.QUARTERLY.value:
            monthly_payment = amount / 3
            frequency = "quarterly"
        elif recurrence == Recurrence.MONTHLY.value:
            monthly_payment = amount
            frequency = "monthly"
        else:
            monthly_payment = amount
            frequency = "one_time"

        if recurrence == Recurrence.NONE.value:
            total_debt = amount
            basis = "One-time amount"
        elif recurrence_mode == RecurrenceMode.DISTRIBUTE.value:
            total_debt = amount
            basis = "Distributed total amount"
        elif recurrence_count and recurrence_count > 0:
            total_debt = amount * recurrence_count
            basis = f"Per-period x {recurrence_count}"
        else:
            total_debt = amount
            basis = "Recurring (no count)"

        debts.append({
            "creditor": flow.get("label", "Unnamed debt"),
            "entity": flow.get("entity") or entity_map.get(flow.get("entity_id", ""), "Unknown"),
            "entity_id": flow.get("entity_id", ""),
            "total_debt_chf": round(total_debt, 2),
            "monthly_payment_chf": round(monthly_payment, 2),
            "frequency": frequency,
            "calculation_basis": basis,
            "source_flow_id": flow.get("id", ""),
        })

    debts.sort(key=lambda d: d["total_debt_chf"], reverse=True)
    return debts

@api_router.put("/treasury/debts/{flow_id}", response_model=CashFlow)
async def update_treasury_debt(flow_id: str, update: TreasuryDebtUpdate):
    """Update debt from treasury context while preserving underlying cash flow behavior."""
    flow = await db.cash_flows.find_one({"id": flow_id}, {"_id": 0})
    if not flow:
        raise HTTPException(status_code=404, detail="Debt flow not found")
    if flow.get("category") != Category.DEBT.value:
        raise HTTPException(status_code=400, detail="Flow is not categorized as debt")

    patch = {}
    if update.creditor is not None:
        creditor = update.creditor.strip()
        if not creditor:
            raise HTTPException(status_code=400, detail="Creditor cannot be empty")
        patch["label"] = creditor

    if update.total_debt_chf is not None:
        total_abs = abs(float(update.total_debt_chf))
        recurrence = flow.get("recurrence", Recurrence.NONE.value)
        recurrence_mode = flow.get("recurrence_mode", RecurrenceMode.REPEAT.value)
        recurrence_count = flow.get("recurrence_count")

        if recurrence == Recurrence.NONE.value:
            signed_amount = -total_abs
        elif recurrence_mode == RecurrenceMode.DISTRIBUTE.value:
            signed_amount = -total_abs
        elif recurrence_count and recurrence_count > 0:
            signed_amount = -(total_abs / recurrence_count)
        else:
            signed_amount = -total_abs
        patch["amount"] = signed_amount

    if not patch:
        raise HTTPException(status_code=400, detail="No valid debt update data")

    return await update_cash_flow(flow_id, CashFlowUpdate(**patch))

# ============== CASH FLOW ROUTES ==============
@api_router.get("/cash-flows")
async def get_cash_flows(entity_id: Optional[str] = None):
    query = {"entity_id": entity_id} if entity_id else {}
    flows = await db.cash_flows.find(query, {"_id": 0}).to_list(1000)
    # Calculate dynamic amounts for percentage-based flows
    result = []
    for f in flows:
        result.append(await get_flow_with_dynamic_amount(f))
    return result

@api_router.get("/cash-flows/with-linked")
async def get_cash_flows_with_linked(entity_id: Optional[str] = None):
    """Get all cash flows grouped by parent-child, with dynamic amount calculation."""
    query = {"entity_id": entity_id} if entity_id else {}
    flows = await db.cash_flows.find(query, {"_id": 0}).to_list(1000)
    
    # Calculate dynamic amounts
    flows_with_amounts = []
    for f in flows:
        flows_with_amounts.append(await get_flow_with_dynamic_amount(f))
    
    # Separate parents and children
    parent_flows = [f for f in flows_with_amounts if not f.get("parent_id")]
    linked_map = {}
    for f in flows_with_amounts:
        if f.get("parent_id"):
            pid = f["parent_id"]
            if pid not in linked_map:
                linked_map[pid] = []
            linked_map[pid].append(f)
    
    result = []
    for pf in parent_flows:
        result.append({
            "flow": pf,
            "linked_flows": linked_map.get(pf["id"], [])
        })
    result.sort(key=lambda x: x["flow"].get("date", ""), reverse=True)
    return result

@api_router.post("/cash-flows", response_model=CashFlow)
async def create_cash_flow(flow: CashFlowCreate):
    entity = await db.entities.find_one({"id": flow.entity_id})
    if not entity:
        raise HTTPException(status_code=400, detail="Entity not found")
    flow_data = flow.model_dump()
    flow_data["amount"] = normalize_amount_for_category(flow_data["category"], flow_data["amount"])
    flow_obj = CashFlow(**flow_data)
    await db.cash_flows.insert_one(flow_obj.model_dump())
    await push_undo("create", "cash_flows", [flow_obj.id], description=f"Create: {flow_obj.label}")
    return flow_obj

@api_router.post("/cash-flows/batch")
async def create_cash_flow_batch(batch: CashFlowBatchCreate):
    """Create parent + linked flows. Children inherit recurrence from parent."""
    entity = await db.entities.find_one({"id": batch.parent.entity_id})
    if not entity:
        raise HTTPException(status_code=400, detail="Entity not found")
    
    # Create parent
    parent_data = batch.parent.model_dump()
    parent_data["amount"] = normalize_amount_for_category(parent_data["category"], parent_data["amount"])
    parent_obj = CashFlow(**parent_data)
    await db.cash_flows.insert_one(parent_obj.model_dump())
    
    # Create linked flows with inherited recurrence
    linked_results = []
    for linked in batch.linked:
        data = linked.model_dump()
        data["parent_id"] = parent_obj.id
        # ENFORCE: Children inherit parent's recurrence and mode
        data["recurrence"] = batch.parent.recurrence
        data["recurrence_mode"] = batch.parent.recurrence_mode
        data["recurrence_end"] = batch.parent.recurrence_end
        data["recurrence_count"] = batch.parent.recurrence_count
        
        # For percentage-based: calculate initial amount from parent
        if data.get("is_percentage") and data.get("percentage_of_parent"):
            pct = data["percentage_of_parent"]
            parent_abs = abs(batch.parent.amount)
            data["amount"] = normalize_amount_for_category(
                data.get("category", Category.EXPENSE.value),
                parent_abs * pct / 100
            )
        else:
            data["amount"] = normalize_amount_for_category(
                data.get("category", Category.EXPENSE.value),
                data.get("amount", 0)
            )
        
        linked_obj = CashFlow(**data)
        await db.cash_flows.insert_one(linked_obj.model_dump())
        linked_results.append(linked_obj.model_dump())
    
    all_ids = [parent_obj.id] + [lr["id"] for lr in linked_results]
    await push_undo("batch_create", "cash_flows", all_ids, description=f"Create: {parent_obj.label} + {len(linked_results)} linked")
    
    return {"parent": parent_obj.model_dump(), "linked": linked_results}
    # Note: undo tracking done above after all inserts

@api_router.put("/cash-flows/{flow_id}", response_model=CashFlow)
async def update_cash_flow(flow_id: str, update: CashFlowUpdate):
    """Update flow. Propagate changes to children for parent flows."""
    update_data = {k: v for k, v in update.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="No valid update data")
    
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    # Get current flow to check if parent — snapshot for undo
    current = await db.cash_flows.find_one({"id": flow_id}, {"_id": 0})
    if not current:
        raise HTTPException(status_code=404, detail="Cash flow not found")

    # Enforce canonical sign behavior on edit:
    # - if amount provided: normalize against effective category
    # - if only category changes: re-normalize existing amount
    effective_category = update_data.get("category", current.get("category", Category.EXPENSE.value))
    if "amount" in update_data:
        update_data["amount"] = normalize_amount_for_category(effective_category, update_data["amount"])
    elif "category" in update_data and current.get("amount") is not None:
        update_data["amount"] = normalize_amount_for_category(update_data["category"], current["amount"])
    
    # Snapshot children too for undo
    children_snapshot = await db.cash_flows.find({"parent_id": flow_id}, {"_id": 0}).to_list(100)
    all_snapshots = [current] + children_snapshot
    
    # Update the flow
    result = await db.cash_flows.find_one_and_update(
        {"id": flow_id}, {"$set": update_data},
        return_document=True, projection={"_id": 0}
    )
    
    # Check if this is a parent flow (has children)
    children = await db.cash_flows.find({"parent_id": flow_id}, {"_id": 0}).to_list(100)
    
    if children:
        child_updates = {"updated_at": datetime.now(timezone.utc).isoformat()}
        
        # PROPAGATE: Recurrence changes to all children
        if any(k in update_data for k in ["recurrence", "recurrence_mode", "recurrence_end", "recurrence_count"]):
            if "recurrence" in update_data:
                child_updates["recurrence"] = update_data["recurrence"]
            if "recurrence_mode" in update_data:
                child_updates["recurrence_mode"] = update_data["recurrence_mode"]
            if "recurrence_end" in update_data:
                child_updates["recurrence_end"] = update_data["recurrence_end"]
            if "recurrence_count" in update_data:
                child_updates["recurrence_count"] = update_data["recurrence_count"]
        
        # PROPAGATE: Date changes to children
        if "date" in update_data:
            child_updates["date"] = update_data["date"]
        
        # PROPAGATE: Certainty changes to children
        if "certainty" in update_data:
            child_updates["certainty"] = update_data["certainty"]
        
        if len(child_updates) > 1:  # More than just updated_at
            await db.cash_flows.update_many(
                {"parent_id": flow_id},
                {"$set": child_updates}
            )
        
        # RECALCULATE: Percentage-based children when parent amount changes
        if "amount" in update_data:
            new_parent_amount = abs(update_data["amount"])
            for child in children:
                if child.get("is_percentage") and child.get("percentage_of_parent"):
                    computed = new_parent_amount * child["percentage_of_parent"] / 100
                    new_child_amount = normalize_amount_for_category(
                        child.get("category", Category.EXPENSE.value),
                        computed
                    )
                    await db.cash_flows.update_one(
                        {"id": child["id"]},
                        {"$set": {"amount": new_child_amount, "updated_at": datetime.now(timezone.utc).isoformat()}}
                    )
    
    await push_undo("update", "cash_flows", [flow_id], previous_data=all_snapshots, description=f"Edit: {current.get('label', '')}")
    
    return result

@api_router.delete("/cash-flows/{flow_id}")
async def delete_cash_flow(flow_id: str, delete_linked: bool = False):
    """Delete flow. Optionally delete linked children or orphan them."""
    # Snapshot for undo — ALWAYS snapshot children for full state restore
    current = await db.cash_flows.find_one({"id": flow_id}, {"_id": 0})
    children = await db.cash_flows.find({"parent_id": flow_id}, {"_id": 0}).to_list(100)
    all_snapshots = ([current] if current else []) + children
    
    if delete_linked:
        await db.cash_flows.delete_many({"parent_id": flow_id})
    else:
        # Orphan children (remove parent_id)
        await db.cash_flows.update_many(
            {"parent_id": flow_id},
            {"$set": {"parent_id": None}}
        )
    
    result = await db.cash_flows.delete_one({"id": flow_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Cash flow not found")
    
    await push_undo(
        "delete", "cash_flows", [flow_id], previous_data=all_snapshots,
        description=f"Delete: {current.get('label', '') if current else 'flow'}",
        side_effects={"delete_linked": delete_linked, "orphaned_child_ids": [c["id"] for c in children] if not delete_linked else []}
    )
    
    return {"message": "Cash flow deleted"}

# ============== FLOW OCCURRENCE ROUTES ==============
@api_router.get("/flow-occurrences")
async def get_flow_occurrences(month: Optional[str] = None, flow_id: Optional[str] = None):
    """Get occurrence records (actuals + variance actions). Filter by month and/or flow_id."""
    query = {}
    if month:
        query["month"] = month
    if flow_id:
        query["flow_id"] = flow_id
    occurrences = await db.flow_occurrences.find(query, {"_id": 0}).to_list(1000)
    return occurrences

async def upsert_flow_occurrence(update: FlowOccurrenceUpdate, suppress_undo: bool = False):
    """Record an actual amount for a flow occurrence.
    
    Variance = planned - actual.
    If variance_action = 'carry_forward': creates a one-time flow for the difference in next month.
    If variance_action = 'write_off': difference is ignored.
    Pushes full undo state including occurrence + carryover flows.
    """
    # Snapshot previous state for undo
    prev_occurrence = await db.flow_occurrences.find_one(
        {"flow_id": update.flow_id, "month": update.month}, {"_id": 0}
    )
    
    # Snapshot existing carryovers that will be deleted
    existing_carryovers = await db.cash_flows.find({
        "carryover_from": update.flow_id,
        "carryover_month": update.month
    }, {"_id": 0}).to_list(100)
    
    # Upsert the occurrence record
    update_fields = {}
    if update.actual_amount is not None:
        update_fields["actual_amount"] = update.actual_amount
    if update.variance_action is not None:
        update_fields["variance_action"] = update.variance_action
    
    if prev_occurrence:
        await db.flow_occurrences.update_one(
            {"flow_id": update.flow_id, "month": update.month},
            {"$set": update_fields}
        )
    else:
        occ = FlowOccurrence(
            flow_id=update.flow_id,
            month=update.month,
            actual_amount=update.actual_amount,
            variance_action=update.variance_action,
        )
        await db.flow_occurrences.insert_one(occ.model_dump())
    
    # Remove any old carryover for this flow+month first
    await db.cash_flows.delete_many({
        "carryover_from": update.flow_id,
        "carryover_month": update.month
    })
    
    # Handle carry_forward: create a flow for the variance in next month
    created_carryover_ids = []
    if update.variance_action == "carry_forward" and update.actual_amount is not None:
        original = await db.cash_flows.find_one({"id": update.flow_id}, {"_id": 0})
        if original:
            # Get planned amount for this occurrence
            planned = original["amount"]
            rec_mode = original.get("recurrence_mode", "repeat")
            rec_count = original.get("recurrence_count")
            if rec_mode == "distribute" and rec_count and rec_count > 0:
                planned = round(planned / rec_count, 2)
            
            variance = planned - update.actual_amount
            
            if abs(variance) > 0.01:
                year, mon = update.month.split("-")
                next_date = date(int(year), int(mon), 1) + relativedelta(months=1)
                next_month_str = next_date.strftime("%Y-%m")
                
                carryover = CashFlow(
                    label=f"{original['label']} (variance carryover)",
                    amount=round(variance, 2),
                    date=f"{next_month_str}-01",
                    category=original.get("category", "Expense"),
                    certainty="Materialized",
                    recurrence="none",
                    recurrence_mode="repeat",
                    entity_id=original.get("entity_id", ""),
                    carryover_from=update.flow_id,
                    carryover_month=update.month,
                )
                await db.cash_flows.insert_one(carryover.model_dump())
                created_carryover_ids.append(carryover.id)
    
    if not suppress_undo:
        # Push undo with full dependency chain
        await push_undo(
            "record_actual", "flow_occurrences",
            [update.flow_id],
            description=f"Actual: {update.flow_id[:8]}... {update.month}",
            side_effects={
                "flow_id": update.flow_id,
                "month": update.month,
                "previous_occurrence": prev_occurrence,
                "deleted_carryovers": existing_carryovers,
                "created_carryover_ids": created_carryover_ids,
            }
        )
    
    # Build response with useful info
    response = {"status": "ok"}
    if created_carryover_ids and update.actual_amount is not None:
        original = await db.cash_flows.find_one({"id": update.flow_id}, {"_id": 0})
        if original:
            resp_planned = original["amount"]
            rec_mode = original.get("recurrence_mode", "repeat")
            rec_count = original.get("recurrence_count")
            if rec_mode == "distribute" and rec_count and rec_count > 0:
                resp_planned = round(resp_planned / rec_count, 2)
            resp_variance = resp_planned - update.actual_amount
            year, mon = update.month.split("-")
            next_date = date(int(year), int(mon), 1) + relativedelta(months=1)
            response["carryover_info"] = {
                "target_month": next_date.strftime("%b %Y"),
                "amount": round(resp_variance, 2),
            }
    response["created_carryover_ids"] = created_carryover_ids
    
    return response

@api_router.put("/flow-occurrences")
async def set_flow_occurrence(update: FlowOccurrenceUpdate):
    return await upsert_flow_occurrence(update, suppress_undo=False)

@api_router.delete("/flow-occurrences")
async def clear_flow_occurrence(flow_id: str, month: str):
    """Clear actual recording for a flow occurrence (revert to planned). Pushes undo."""
    # Snapshot for undo
    prev_occurrence = await db.flow_occurrences.find_one(
        {"flow_id": flow_id, "month": month}, {"_id": 0}
    )
    existing_carryovers = await db.cash_flows.find({
        "carryover_from": flow_id,
        "carryover_month": month
    }, {"_id": 0}).to_list(100)
    
    await db.flow_occurrences.delete_one({"flow_id": flow_id, "month": month})
    await db.cash_flows.delete_many({
        "carryover_from": flow_id,
        "carryover_month": month
    })
    
    # Push undo to restore previous state
    if prev_occurrence:
        await push_undo(
            "record_actual", "flow_occurrences",
            [flow_id],
            description=f"Clear actual: {flow_id[:8]}... {month}",
            side_effects={
                "flow_id": flow_id,
                "month": month,
                "previous_occurrence": prev_occurrence,
                "deleted_carryovers": existing_carryovers,
                "created_carryover_ids": [],
            }
        )
    
    return {"status": "ok"}

async def recalculate_import_batch_counts(batch_id: str):
    rows = await db.actual_import_rows.find({"batch_id": batch_id}, {"_id": 0, "status": 1, "include": 1}).to_list(100000)
    total_rows = len(rows)
    ready_rows = sum(1 for r in rows if r.get("status") == "ready" and r.get("include", True))
    warning_rows = sum(1 for r in rows if r.get("status") == "warning" and r.get("include", True))
    discarded_rows = sum(1 for r in rows if not r.get("include", True) or r.get("status") == "discarded")
    applied_rows = sum(1 for r in rows if r.get("status") == "applied")
    failed_rows = sum(1 for r in rows if r.get("status") == "failed")
    skipped_rows = sum(1 for r in rows if r.get("status") == "skipped")

    await db.actual_import_batches.update_one(
        {"id": batch_id},
        {"$set": {
            "total_rows": total_rows,
            "ready_rows": ready_rows,
            "warning_rows": warning_rows,
            "discarded_rows": discarded_rows,
            "applied_rows": applied_rows,
            "failed_rows": failed_rows,
            "skipped_rows": skipped_rows,
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }},
    )

@api_router.post("/actual-imports/parse")
async def parse_actual_import(
    file: UploadFile = File(...),
    entity_id: Optional[str] = Form(None),
    user: dict = Depends(get_current_user),
):
    ensure_bulk_actuals_enabled()
    filename = file.filename or "upload"
    ext = Path(filename).suffix.lower()
    if ext not in [".csv", ".xlsx"]:
        raise HTTPException(status_code=400, detail="Only CSV/XLSX files are supported in v1")

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")

    try:
        if ext == ".csv":
            parsed_columns, parsed_rows = parse_csv_to_rows(content)
        else:
            parsed_columns: List[str] = []
            parsed_rows: List[Dict[str, Any]] = []
            used_openpyxl = False
            try:
                from openpyxl import load_workbook

                workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
                sheet = workbook.active
                rows_iter = sheet.iter_rows(values_only=True)
                header_row = next(rows_iter, None)
                if not header_row:
                    raise ValueError("No header row found in XLSX")
                parsed_columns = [str(h).strip() if h is not None else "" for h in header_row]
                for values in rows_iter:
                    row_dict: Dict[str, Any] = {}
                    for i, col in enumerate(parsed_columns):
                        key = col or f"column_{i+1}"
                        row_dict[key] = values[i] if i < len(values) else None
                    parsed_rows.append(row_dict)
                used_openpyxl = True
            except Exception:
                used_openpyxl = False
            if not used_openpyxl:
                try:
                    parsed_columns, parsed_rows = read_first_sheet_as_dict_rows(content)
                except Exception as exc:
                    raise HTTPException(
                        status_code=400,
                        detail=f"Could not read XLSX ({exc}). Redeploy with openpyxl if this persists: pip install openpyxl==3.1.5",
                    ) from exc
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Could not parse file: {exc}")

    if not parsed_rows:
        raise HTTPException(status_code=400, detail="No rows found in file")
    if len(parsed_rows) > 10000:
        raise HTTPException(status_code=400, detail="File too large for v1 import (max 10,000 rows)")

    detected = detect_import_columns(parsed_columns)
    if "date" not in detected or not detected.get("date") or "description" not in detected or not detected.get("description"):
        preview = ", ".join(str(c) for c in parsed_columns[:12]) if parsed_columns else "(no columns)"
        raise HTTPException(
            status_code=400,
            detail=f"Could not detect required date/description columns. Found headers: {preview}",
        )
    if ("amount" not in detected or not detected.get("amount")) and not (
        detected.get("debit") and detected.get("credit")
    ):
        preview = ", ".join(str(c) for c in parsed_columns[:12]) if parsed_columns else "(no columns)"
        raise HTTPException(
            status_code=400,
            detail=f"Could not detect amount column (or debit+credit pair). Found headers: {preview}",
        )

    if entity_id:
        entity_exists = await db.entities.find_one({"id": entity_id}, {"_id": 0, "id": 1})
        if not entity_exists:
            raise HTTPException(status_code=400, detail="entity_id does not exist")

    flow_query = {"entity_id": entity_id} if entity_id else {}
    flows = await db.cash_flows.find(flow_query, {"_id": 0}).to_list(5000)

    batch = ActualImportBatch(
        filename=filename,
        file_type=ext.lstrip("."),
        entity_id=entity_id,
    )
    await db.actual_import_batches.insert_one(batch.model_dump())

    created_rows = []
    for idx, row in enumerate(parsed_rows):
        raw_date = row.get(detected["date"])
        parsed_d = parse_import_row_date(raw_date)
        if not parsed_d:
            continue
        transaction_date = parsed_d.strftime("%Y-%m-%d")
        month = parsed_d.strftime("%Y-%m")

        description = str(row.get(detected["description"], "") or "").strip()
        if not description:
            description = f"Imported row {idx + 1}"

        if "amount" in detected:
            parsed_amount = parse_import_amount(row.get(detected["amount"]))
        else:
            debit = parse_import_amount(row.get(detected["debit"])) or 0
            credit = parse_import_amount(row.get(detected["credit"])) or 0
            parsed_amount = credit - debit

        if parsed_amount is None or parsed_amount == 0:
            continue

        category = Category.REVENUE if parsed_amount > 0 else Category.EXPENSE
        match = best_flow_match(flows, description, parsed_amount, entity_id)
        selected_flow_id = match["flow_id"] if match["score"] >= 0.7 else None
        matched_entity_id = entity_id
        if not matched_entity_id and selected_flow_id:
            matched = next((f for f in flows if f.get("id") == selected_flow_id), None)
            matched_entity_id = matched.get("entity_id") if matched else None
        include = True
        status = "ready" if selected_flow_id else "warning"

        import_row = ActualImportRow(
            batch_id=batch.id,
            row_index=int(idx),
            include=include,
            status=status,
            raw_date=str(raw_date),
            raw_description=description,
            raw_amount=str(row.get(detected.get("amount", ""), "")),
            transaction_date=transaction_date,
            month=month,
            description=description,
            amount=float(parsed_amount),
            category=category,
            entity_id=matched_entity_id,
            suggested_flow_id=match["flow_id"],
            selected_flow_id=selected_flow_id,
            match_score=match["score"],
            match_reason=match["reason"],
            classification="existing_flow",
            actual_merge_mode="override",
        )
        created_rows.append(import_row.model_dump())

    if not created_rows:
        await db.actual_import_batches.delete_one({"id": batch.id})
        raise HTTPException(status_code=400, detail="No valid transaction rows detected")

    await db.actual_import_rows.insert_many(created_rows)
    await recalculate_import_batch_counts(batch.id)
    batch_doc = await db.actual_import_batches.find_one({"id": batch.id}, {"_id": 0})
    rows_doc = await db.actual_import_rows.find({"batch_id": batch.id}, {"_id": 0}).sort("row_index", 1).to_list(5000)
    return {"batch": batch_doc, "rows": rows_doc, "detected_columns": detected}

@api_router.get("/actual-imports/{batch_id}")
async def get_actual_import_batch(batch_id: str, user: dict = Depends(get_current_user)):
    ensure_bulk_actuals_enabled()
    batch = await db.actual_import_batches.find_one({"id": batch_id}, {"_id": 0})
    if not batch:
        raise HTTPException(status_code=404, detail="Import batch not found")
    return batch

@api_router.get("/actual-imports")
async def list_actual_import_batches(
    entity_id: Optional[str] = None,
    status: Optional[str] = None,
    limit: int = Query(default=30, ge=1, le=200),
    user: dict = Depends(get_current_user),
):
    ensure_bulk_actuals_enabled()
    query: Dict[str, Any] = {}
    if entity_id:
        query["entity_id"] = entity_id
    if status:
        query["status"] = status
    batches = await db.actual_import_batches.find(query, {"_id": 0}).sort("created_at", -1).limit(limit).to_list(limit)
    return batches

@api_router.get("/actual-imports/{batch_id}/rows")
async def get_actual_import_rows(batch_id: str, user: dict = Depends(get_current_user)):
    ensure_bulk_actuals_enabled()
    rows = await db.actual_import_rows.find({"batch_id": batch_id}, {"_id": 0}).sort("row_index", 1).to_list(100000)
    return rows

@api_router.put("/actual-imports/{batch_id}/rows/{row_id}")
async def update_actual_import_row(batch_id: str, row_id: str, update: ActualImportRowUpdate, user: dict = Depends(get_current_user)):
    ensure_bulk_actuals_enabled()
    batch = await db.actual_import_batches.find_one({"id": batch_id}, {"_id": 0})
    if not batch:
        raise HTTPException(status_code=404, detail="Import batch not found")
    row = await db.actual_import_rows.find_one({"id": row_id, "batch_id": batch_id}, {"_id": 0})
    if not row:
        raise HTTPException(status_code=404, detail="Import row not found")

    patch = {k: v for k, v in update.model_dump().items() if v is not None}
    if "transaction_date" in patch:
        normalized_date = normalize_import_transaction_date(patch["transaction_date"])
        if not normalized_date:
            raise HTTPException(status_code=400, detail="Invalid transaction_date format")
        patch["transaction_date"] = normalized_date
        patch.setdefault("month", normalized_date[:7])
    if "month" in patch and not is_valid_month_key(str(patch["month"])):
        raise HTTPException(status_code=400, detail="Invalid month format, expected YYYY-MM")
    if "amount" in patch:
        try:
            amount_value = float(patch["amount"])
        except Exception:
            raise HTTPException(status_code=400, detail="Invalid amount")
        if not math.isfinite(amount_value):
            raise HTTPException(status_code=400, detail="Invalid amount")
        patch["amount"] = amount_value
    if "description" in patch:
        patch["description"] = str(patch["description"]).strip()
        if not patch["description"]:
            raise HTTPException(status_code=400, detail="Description cannot be empty")
    if "entity_id" in patch and patch["entity_id"]:
        entity_exists = await db.entities.find_one({"id": patch["entity_id"]}, {"_id": 0, "id": 1})
        if not entity_exists:
            raise HTTPException(status_code=400, detail="entity_id does not exist")

    if "variance_action" in patch and patch["variance_action"] not in {"actual_only", "carry_forward", "write_off"}:
        raise HTTPException(status_code=400, detail="Invalid variance_action")
    if "classification" in patch and patch["classification"] not in {"existing_flow", "new_flow"}:
        raise HTTPException(status_code=400, detail="Invalid classification")
    if "actual_merge_mode" in patch and patch["actual_merge_mode"] not in {"override", "addition"}:
        raise HTTPException(status_code=400, detail="Invalid actual_merge_mode")

    if patch.get("classification") == "new_flow":
        patch["selected_flow_id"] = None

    merged_class = patch["classification"] if "classification" in patch else row.get("classification", "existing_flow")
    row_entity_id = patch.get("entity_id", row.get("entity_id"))
    batch_entity_id = batch.get("entity_id")
    sel_for_validate = None if merged_class == "new_flow" else (
        patch["selected_flow_id"] if "selected_flow_id" in patch else row.get("selected_flow_id")
    )
    if sel_for_validate:
        await validate_selected_flow_for_row(sel_for_validate, row_entity_id, batch_entity_id)

    merged_include = patch["include"] if "include" in patch else row.get("include", True)
    ent_ok = bool(row_entity_id or batch_entity_id)
    if not merged_include:
        patch["status"] = "discarded"
    elif merged_class == "new_flow":
        patch["status"] = "ready" if ent_ok else "warning"
    elif sel_for_validate:
        patch["status"] = "ready"
    else:
        patch["status"] = "warning"

    patch["updated_at"] = datetime.now(timezone.utc).isoformat()

    updated = await db.actual_import_rows.find_one_and_update(
        {"id": row_id, "batch_id": batch_id},
        {"$set": patch},
        return_document=True,
        projection={"_id": 0},
    )
    await recalculate_import_batch_counts(batch_id)
    return updated

@api_router.post("/actual-imports/{batch_id}/apply")
async def apply_actual_import(batch_id: str, payload: ActualImportApplyRequest, user: dict = Depends(get_current_user)):
    ensure_bulk_actuals_enabled()
    batch = await db.actual_import_batches.find_one({"id": batch_id}, {"_id": 0})
    if not batch:
        raise HTTPException(status_code=404, detail="Import batch not found")
    if batch.get("status") == "discarded":
        raise HTTPException(status_code=400, detail="Import batch was discarded")

    rows = await db.actual_import_rows.find({"batch_id": batch_id}, {"_id": 0}).sort("row_index", 1).to_list(100000)
    fingerprint = build_apply_fingerprint(rows, payload.idempotency_key)

    if batch.get("last_apply_fingerprint") == fingerprint and batch.get("status") in {"applied", "partial"}:
        return {
            "batch_id": batch_id,
            "status": "idempotent",
            "applied_rows": batch.get("applied_rows", 0),
            "failed_rows": batch.get("failed_rows", 0),
            "skipped_rows": batch.get("skipped_rows", 0),
            "discarded_rows": batch.get("discarded_rows", 0),
            "errors": [],
        }
    if batch.get("status") == "applied" and batch.get("last_apply_fingerprint") and batch.get("last_apply_fingerprint") != fingerprint:
        raise HTTPException(status_code=409, detail="Batch already applied. Create a new import batch to apply different rows.")

    to_apply = [r for r in rows if r.get("include", True)]
    if not to_apply:
        raise HTTPException(status_code=400, detail="No included rows to apply")
    if len(to_apply) > 5000:
        raise HTTPException(status_code=400, detail="Too many included rows to apply at once (max 5,000)")

    applied = 0
    failed = 0
    skipped = 0
    discarded = len([r for r in rows if not r.get("include", True)])
    errors = []
    batch_side_effects = []

    # Several bank lines may map to the same flow + month; each row has its own actual_merge_mode
    # (Replace vs Add to current). Rows are applied in file order.

    for row in to_apply:
        row_id = row["id"]
        classification = row.get("classification", "existing_flow")
        month = row.get("month")
        amount = row.get("amount")
        variance_action = row.get("variance_action", "actual_only")
        entity_for_row = row.get("entity_id") or batch.get("entity_id")

        if not month or amount is None:
            failed += 1
            msg = "Row missing month or amount"
            errors.append({"row_id": row_id, "error": msg})
            await db.actual_import_rows.update_one(
                {"id": row_id},
                {"$set": {"status": "failed", "error": msg, "updated_at": datetime.now(timezone.utc).isoformat()}},
            )
            continue
        if not is_valid_month_key(str(month)):
            failed += 1
            msg = "Row has invalid month format (expected YYYY-MM)"
            errors.append({"row_id": row_id, "error": msg})
            await db.actual_import_rows.update_one(
                {"id": row_id},
                {"$set": {"status": "failed", "error": msg, "updated_at": datetime.now(timezone.utc).isoformat()}},
            )
            continue
        try:
            amount_value = float(amount)
        except Exception:
            amount_value = float("nan")
        if not math.isfinite(amount_value):
            failed += 1
            msg = "Row has invalid amount"
            errors.append({"row_id": row_id, "error": msg})
            await db.actual_import_rows.update_one(
                {"id": row_id},
                {"$set": {"status": "failed", "error": msg, "updated_at": datetime.now(timezone.utc).isoformat()}},
            )
            continue

        flow_id: Optional[str] = row.get("selected_flow_id")
        if classification == "new_flow":
            if not entity_for_row:
                failed += 1
                msg = "New line rows require entity scope (set entity on batch or row)"
                errors.append({"row_id": row_id, "error": msg})
                await db.actual_import_rows.update_one(
                    {"id": row_id},
                    {"$set": {"status": "failed", "error": msg, "updated_at": datetime.now(timezone.utc).isoformat()}},
                )
                continue
            flow_id = None
        elif not flow_id:
            failed += 1
            msg = "Row missing selected_flow_id (pick a flow or switch target to New line)"
            errors.append({"row_id": row_id, "error": msg})
            await db.actual_import_rows.update_one(
                {"id": row_id},
                {"$set": {"status": "failed", "error": msg, "updated_at": datetime.now(timezone.utc).isoformat()}},
            )
            continue

        created_flow_id: Optional[str] = None
        try:
            if classification == "new_flow":
                flow_obj = build_cash_flow_from_import_row(row, entity_for_row, amount_value)
                await db.cash_flows.insert_one(flow_obj.model_dump())
                flow_id = flow_obj.id
                created_flow_id = flow_id
                await validate_selected_flow_for_row(flow_id, entity_for_row, batch.get("entity_id"))
            else:
                await validate_selected_flow_for_row(flow_id, row.get("entity_id"), batch.get("entity_id"))

            desired_variance_action = None if variance_action == "actual_only" else variance_action
            prev_occurrence = await db.flow_occurrences.find_one(
                {"flow_id": flow_id, "month": month},
                {"_id": 0},
            )
            existing_amount = prev_occurrence.get("actual_amount") if prev_occurrence else None
            existing_variance_action = prev_occurrence.get("variance_action") if prev_occurrence else None

            base_existing = float(existing_amount) if existing_amount is not None else 0.0
            row_merge = row.get("actual_merge_mode") or payload.actual_merge_mode
            if row_merge == "addition":
                final_amount = base_existing + amount_value
            else:
                final_amount = amount_value

            if (
                prev_occurrence
                and existing_amount is not None
                and round(float(existing_amount), 2) == round(float(final_amount), 2)
                and existing_variance_action == desired_variance_action
            ):
                skipped += 1
                if created_flow_id:
                    await db.cash_flows.delete_many({"parent_id": created_flow_id})
                    await db.cash_flows.delete_one({"id": created_flow_id})
                await db.actual_import_rows.update_one(
                    {"id": row_id},
                    {"$set": {"status": "skipped", "error": None, "updated_at": datetime.now(timezone.utc).isoformat()}},
                )
                continue

            existing_carryovers = await db.cash_flows.find({
                "carryover_from": flow_id,
                "carryover_month": month
            }, {"_id": 0}).to_list(100)

            occ_update = FlowOccurrenceUpdate(
                flow_id=flow_id,
                month=month,
                actual_amount=float(final_amount),
                variance_action=desired_variance_action,
            )
            occ_result = await upsert_flow_occurrence(occ_update, suppress_undo=True)
            applied += 1
            row_effects: Dict[str, Any] = {
                "flow_id": flow_id,
                "month": month,
                "previous_occurrence": prev_occurrence,
                "deleted_carryovers": existing_carryovers,
                "created_carryover_ids": occ_result.get("created_carryover_ids", []),
            }
            if created_flow_id:
                row_effects["created_flow_id"] = created_flow_id
            batch_side_effects.append(row_effects)
            row_set: Dict[str, Any] = {
                "status": "applied",
                "error": None,
                "updated_at": datetime.now(timezone.utc).isoformat(),
            }
            if classification == "new_flow":
                row_set["selected_flow_id"] = flow_id
            await db.actual_import_rows.update_one({"id": row_id}, {"$set": row_set})
        except Exception as exc:
            if created_flow_id:
                await db.cash_flows.delete_many({"parent_id": created_flow_id})
                await db.cash_flows.delete_one({"id": created_flow_id})
            failed += 1
            msg = str(exc)
            errors.append({"row_id": row_id, "error": msg})
            await db.actual_import_rows.update_one(
                {"id": row_id},
                {"$set": {"status": "failed", "error": msg, "updated_at": datetime.now(timezone.utc).isoformat()}},
            )

    if applied > 0:
        await push_undo(
            "record_actual_batch",
            "flow_occurrences",
            list({s["flow_id"] for s in batch_side_effects}),
            description=f"Bulk actual import ({applied} rows)",
            side_effects={"batch_id": batch_id, "rows": batch_side_effects},
        )

    final_status = "applied" if failed == 0 else ("partial" if (applied > 0 or skipped > 0) else "failed")
    await db.actual_import_batches.update_one(
        {"id": batch_id},
        {"$set": {
            "status": final_status,
            "last_apply_fingerprint": fingerprint,
            "last_applied_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }},
    )
    await recalculate_import_batch_counts(batch_id)

    return {
        "batch_id": batch_id,
        "status": final_status,
        "applied_rows": applied,
        "failed_rows": failed,
        "skipped_rows": skipped,
        "discarded_rows": discarded,
        "duplicate_conflicts": 0,
        "errors": errors[:50],
    }

@api_router.post("/actual-imports/{batch_id}/discard")
async def discard_actual_import(batch_id: str, user: dict = Depends(get_current_user)):
    ensure_bulk_actuals_enabled()
    result = await db.actual_import_batches.update_one(
        {"id": batch_id},
        {"$set": {"status": "discarded", "updated_at": datetime.now(timezone.utc).isoformat()}},
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Import batch not found")
    return {"status": "discarded"}

# ============== SETTINGS ROUTES ==============
@api_router.get("/settings", response_model=Settings)
async def get_settings():
    settings = await db.settings.find_one({"id": "settings"}, {"_id": 0})
    if not settings:
        default = Settings()
        await db.settings.insert_one(default.model_dump())
        return default
    return settings

@api_router.put("/settings", response_model=Settings)
async def update_settings(update: SettingsUpdate):
    result = await db.settings.find_one_and_update(
        {"id": "settings"},
        {"$set": {"safety_buffer": update.safety_buffer}},
        upsert=True, return_document=True, projection={"_id": 0}
    )
    return result

@api_router.get("/meta/cash-flow")
async def get_cash_flow_meta():
    return {
        "categories": [c.value for c in Category],
        "variance_actions": [
            {"value": "actual_only", "label": "Actual only"},
            {"value": "carry_forward", "label": "Carry delta forward"},
            {"value": "write_off", "label": "Write off delta"},
        ],
    }

# ============== PROJECTION ENGINE ==============
def expand_recurring_flows(flows: List[dict], start_date: date, end_date: date) -> List[dict]:
    """Expand recurring flows. Supports monthly/quarterly recurrence and repeat/distribute modes."""
    expanded = []
    
    for flow in flows:
        flow_date = date.fromisoformat(flow["date"][:10])
        flow_id = flow.get("id", "")
        recurrence = flow.get("recurrence", "none")
        recurrence_mode = flow.get("recurrence_mode", "repeat")
        
        if recurrence == "none":
            if start_date <= flow_date <= end_date:
                expanded.append({
                    "date": flow_date,
                    "amount": flow["amount"],
                    "label": flow["label"],
                    "certainty": flow["certainty"],
                    "category": flow["category"],
                    "entity_id": flow.get("entity_id", ""),
                    "flow_id": flow_id,
                    "parent_id": flow.get("parent_id"),
                    "is_percentage": flow.get("is_percentage", False),
                    "carryover_from": flow.get("carryover_from"),
                })
        else:
            # Monthly or quarterly
            interval_months = 1 if recurrence == "monthly" else 3
            current = flow_date
            count = 0
            max_count = flow.get("recurrence_count") or 999
            end_rec = date.fromisoformat(flow["recurrence_end"][:10]) if flow.get("recurrence_end") else end_date
            
            # Calculate distribute amounts if needed
            total_amount = flow["amount"]
            if recurrence_mode == "distribute" and max_count < 999:
                per_period = round(total_amount / max_count, 2)
                last_period = round(total_amount - per_period * (max_count - 1), 2)
            else:
                per_period = total_amount
                last_period = total_amount
            
            while current <= min(end_date, end_rec) and count < max_count:
                if current >= start_date:
                    # For distribute: last period gets rounding remainder
                    if recurrence_mode == "distribute" and max_count < 999:
                        period_amount = last_period if count == max_count - 1 else per_period
                    else:
                        period_amount = total_amount
                    
                    expanded.append({
                        "date": current,
                        "amount": period_amount,
                        "label": flow["label"],
                        "certainty": flow["certainty"],
                        "category": flow["category"],
                        "entity_id": flow.get("entity_id", ""),
                        "flow_id": flow_id,
                        "parent_id": flow.get("parent_id"),
                        "is_percentage": flow.get("is_percentage", False),
                    })
                count += 1
                current = flow_date + relativedelta(months=count * interval_months)
    
    return expanded

def get_certainty_levels(scenario: str) -> List[str]:
    return {
        "committed": ["Materialized"],
        "likely": ["Materialized", "Sure to happen"],
        "extended": ["Materialized", "Sure to happen", "50/50"],
        "full": ["Materialized", "Sure to happen", "50/50", "Idea"]
    }.get(scenario, ["Materialized"])

@api_router.get("/projection")
async def get_projection(
    scenario: str = "likely",
    entity_id: Optional[str] = Query(None, description="Filter by entity"),
    horizon: int = Query(12, description="Projection horizon in months (12, 24, or 36)")
):
    """Single projection engine - all UI components use this."""
    # Validate horizon
    if horizon not in [12, 24, 36]:
        horizon = 12
    
    # Get bank accounts (state)
    acc_query = {"entity_id": entity_id} if entity_id else {}
    accounts = await db.bank_accounts.find(acc_query, {"_id": 0}).to_list(100)
    cash_now = sum(acc.get("amount", 0) for acc in accounts)
    
    # Get cash flows (events)
    flow_query = {"entity_id": entity_id} if entity_id else {}
    flows = await db.cash_flows.find(flow_query, {"_id": 0}).to_list(1000)
    
    # Calculate dynamic amounts for percentage-based flows
    flows_with_amounts = []
    for f in flows:
        flows_with_amounts.append(await get_flow_with_dynamic_amount(f))
    
    # Get settings
    settings = await db.settings.find_one({"id": "settings"}, {"_id": 0})
    safety_buffer = settings.get("safety_buffer", 50000) if settings else 50000
    
    # Projection based on horizon
    today = date.today()
    start_of_month = today.replace(day=1)
    end_date = start_of_month + relativedelta(months=horizon)
    
    # Filter by certainty
    certainty_levels = get_certainty_levels(scenario)
    filtered = [f for f in flows_with_amounts if f.get("certainty") in certainty_levels]
    
    # Get actuals — when recorded, projection uses actual amount instead of planned
    all_occs = await db.flow_occurrences.find({}, {"_id": 0}).to_list(1000)
    actuals_map = {}  # (flow_id, month) -> actual_amount
    for occ in all_occs:
        if occ.get("actual_amount") is not None:
            actuals_map[(occ["flow_id"], occ["month"])] = occ["actual_amount"]
    
    # Expand recurring
    expanded = expand_recurring_flows(filtered, start_of_month, end_date)
    
    # Apply actuals overlay: replace planned with actual where recorded
    for ef in expanded:
        fid = ef.get("flow_id", "")
        mkey = ef["date"].strftime("%Y-%m")
        if (fid, mkey) in actuals_map:
            ef["amount"] = actuals_map[(fid, mkey)]
    
    # Group by month
    months_data = {}
    for i in range(horizon):
        m = start_of_month + relativedelta(months=i)
        key = m.strftime("%Y-%m")
        months_data[key] = {
            "month": key,
            "month_label": m.strftime("%b %Y"),
            "inflows": 0.0,
            "outflows": 0.0
        }
    
    for flow in expanded:
        key = flow["date"].strftime("%Y-%m")
        if key in months_data:
            if flow["amount"] > 0:
                months_data[key]["inflows"] += flow["amount"]
            else:
                months_data[key]["outflows"] += abs(flow["amount"])
    
    # Calculate projection
    months = []
    closing = cash_now
    lowest_cash = cash_now
    lowest_cash_month = start_of_month.strftime("%b %Y")
    highest_pressure_month = start_of_month.strftime("%b %Y")
    highest_pressure = 0.0
    first_watch_month = None
    first_danger_month = None
    
    for key in sorted(months_data.keys()):
        data = months_data[key]
        net = data["inflows"] - data["outflows"]
        closing += net
        
        if closing >= safety_buffer:
            status = "Good"
        elif closing >= 0:
            status = "Watch"
            if first_watch_month is None:
                first_watch_month = data["month_label"]
        else:
            status = "Danger"
            if first_danger_month is None:
                first_danger_month = data["month_label"]
        
        if closing < lowest_cash:
            lowest_cash = closing
            lowest_cash_month = data["month_label"]
        
        if data["outflows"] > highest_pressure:
            highest_pressure = data["outflows"]
            highest_pressure_month = data["month_label"]
        
        months.append(MonthProjection(
            month=data["month"],
            month_label=data["month_label"],
            inflows=round(data["inflows"], 2),
            outflows=round(data["outflows"], 2),
            net=round(net, 2),
            closing_cash=round(closing, 2),
            status=status
        ))
    
    overall_status = "Danger" if lowest_cash < 0 else ("Watch" if lowest_cash < safety_buffer else "Good")
    
    return ProjectionResponse(
        cash_now=round(cash_now, 2),
        lowest_cash=round(lowest_cash, 2),
        lowest_cash_month=lowest_cash_month,
        highest_pressure_month=highest_pressure_month,
        first_watch_month=first_watch_month,
        first_danger_month=first_danger_month,
        overall_status=overall_status,
        safety_buffer=safety_buffer,
        months=months
    )


@api_router.get("/projection/matrix")
async def get_projection_matrix(
    scenario: str = "likely",
    entity_id: Optional[str] = Query(None),
    horizon: int = Query(12)
):
    """Matrix data derived from the SAME projection engine. No independent computation."""
    if horizon not in [12, 24, 36]:
        horizon = 12
    
    # Same flow fetching as main projection
    flow_query = {"entity_id": entity_id} if entity_id else {}
    flows = await db.cash_flows.find(flow_query, {"_id": 0}).to_list(1000)
    
    flows_with_amounts = []
    for f in flows:
        flows_with_amounts.append(await get_flow_with_dynamic_amount(f))
    
    certainty_levels = get_certainty_levels(scenario)
    filtered = [f for f in flows_with_amounts if f.get("certainty") in certainty_levels]
    
    # Same actuals handling as main projection
    all_occs = await db.flow_occurrences.find({}, {"_id": 0}).to_list(1000)
    actuals_map = {}
    for occ in all_occs:
        if occ.get("actual_amount") is not None:
            actuals_map[(occ["flow_id"], occ["month"])] = occ["actual_amount"]
    
    today = date.today()
    start_of_month = today.replace(day=1)
    end_date = start_of_month + relativedelta(months=horizon)
    
    # Same expand function
    expanded = expand_recurring_flows(filtered, start_of_month, end_date)
    
    # Store planned amounts before actuals overlay
    planned_map = {}  # (flow_id, month) -> planned_amount
    for ef in expanded:
        fid = ef.get("flow_id", "")
        mkey = ef["date"].strftime("%Y-%m")
        if fid:
            planned_map[(fid, mkey)] = ef["amount"]
    
    # Apply actuals overlay
    for ef in expanded:
        fid = ef.get("flow_id", "")
        mkey = ef["date"].strftime("%Y-%m")
        if (fid, mkey) in actuals_map:
            ef["amount"] = actuals_map[(fid, mkey)]
    
    # Build month keys
    month_keys = []
    for i in range(horizon):
        m = start_of_month + relativedelta(months=i)
        month_keys.append({
            "key": m.strftime("%Y-%m"),
            "label": m.strftime("%b %Y")
        })
    
    # Group expanded flows by flow_id + month
    flow_month_map = {}
    flow_info = {}
    
    for ef in expanded:
        fid = ef.get("flow_id", "")
        if not fid:
            continue
        mkey = ef["date"].strftime("%Y-%m")
        
        if fid not in flow_month_map:
            flow_month_map[fid] = {}
            # Use planned amount to determine revenue/expense classification
            planned_amt = planned_map.get((fid, mkey), ef["amount"])
            flow_info[fid] = {
                "flow_id": fid,
                "label": ef["label"],
                "category": ef["category"],
                "is_revenue": planned_amt > 0,
                "parent_id": ef.get("parent_id"),
                "is_percentage": ef.get("is_percentage", False),
            }
        
        flow_month_map[fid][mkey] = flow_month_map[fid].get(mkey, 0) + ef["amount"]
    
    # Build rows: separate revenues and expenses, exclude children (show net on parent)
    revenue_rows = []
    expense_rows = []
    
    # Build a priority lookup from raw flows (NOT from engine, pure metadata)
    flow_priority_map = {f.get("id", ""): f.get("priority") for f in flows}
    
    for fid, info in flow_info.items():
        cells = {}
        for mk in month_keys:
            amt = flow_month_map[fid].get(mk["key"])
            if amt is not None:
                planned = planned_map.get((fid, mk["key"]))
                actual = actuals_map.get((fid, mk["key"]))
                cell_data = {"amount": round(amt, 2)}
                if actual is not None:
                    cell_data["actual"] = round(actual, 2)
                    cell_data["planned"] = round(planned, 2) if planned is not None else round(amt, 2)
                    cell_data["has_actual"] = True
                else:
                    cell_data["has_actual"] = False
                cells[mk["key"]] = cell_data
        
        row = {
            "flow_id": fid,
            "label": info["label"],
            "category": info["category"],
            "parent_id": info["parent_id"],
            "is_percentage": info["is_percentage"],
            "cells": cells,
            "row_total": round(sum(c.get("amount", 0) for c in cells.values()), 2),
            "priority": flow_priority_map.get(fid),
        }
        
        if info["is_revenue"]:
            revenue_rows.append(row)
        else:
            expense_rows.append(row)
    
    # Net per month (from same expanded data — guaranteed match)
    net_per_month = {}
    revenue_per_month = {}
    cost_per_month = {}
    for mk in month_keys:
        net_per_month[mk["key"]] = 0.0
        revenue_per_month[mk["key"]] = 0.0
        cost_per_month[mk["key"]] = 0.0
    for ef in expanded:
        mkey = ef["date"].strftime("%Y-%m")
        if mkey in net_per_month:
            net_per_month[mkey] += ef["amount"]
            if ef["amount"] > 0:
                revenue_per_month[mkey] += ef["amount"]
            else:
                cost_per_month[mkey] += abs(ef["amount"])
    net_per_month = {k: round(v, 2) for k, v in net_per_month.items()}
    revenue_per_month = {k: round(v, 2) for k, v in revenue_per_month.items()}
    cost_per_month = {k: round(v, 2) for k, v in cost_per_month.items()}
    
    # Cash balance per month (starting cash + cumulative net — engine-driven)
    acc_query = {"entity_id": entity_id} if entity_id else {}
    accounts = await db.bank_accounts.find(acc_query, {"_id": 0}).to_list(100)
    cash_now = sum(acc.get("amount", 0) for acc in accounts)
    
    cash_balance_per_month = {}
    running = cash_now
    for mk in month_keys:
        running += net_per_month.get(mk["key"], 0)
        cash_balance_per_month[mk["key"]] = round(running, 2)
    
    # Horizon totals — computed from engine data, never by frontend
    total_revenue = round(sum(revenue_per_month.values()), 2)
    total_cost = round(sum(cost_per_month.values()), 2)
    total_net = round(sum(net_per_month.values()), 2)
    
    return {
        "months": month_keys,
        "revenue_rows": revenue_rows,
        "expense_rows": expense_rows,
        "net_per_month": net_per_month,
        "revenue_per_month": revenue_per_month,
        "cost_per_month": cost_per_month,
        "cash_balance_per_month": cash_balance_per_month,
        "cash_now": round(cash_now, 2),
        "total_revenue": total_revenue,
        "total_cost": total_cost,
        "total_net": total_net,
    }


@api_router.get("/projection/drivers")
async def get_negative_month_drivers(
    scenario: str = "likely",
    entity_id: Optional[str] = Query(None),
    horizon: int = Query(12)
):
    """Top drivers of negative months. Uses projection engine values only. Aggregates by flow label."""
    if horizon not in [12, 24, 36]:
        horizon = 12
    
    flow_query = {"entity_id": entity_id} if entity_id else {}
    flows = await db.cash_flows.find(flow_query, {"_id": 0}).to_list(1000)
    flows_with_amounts = []
    for f in flows:
        flows_with_amounts.append(await get_flow_with_dynamic_amount(f))
    
    certainty_levels = get_certainty_levels(scenario)
    filtered = [f for f in flows_with_amounts if f.get("certainty") in certainty_levels]
    
    all_occs = await db.flow_occurrences.find({}, {"_id": 0}).to_list(1000)
    actuals_map = {}
    for occ in all_occs:
        if occ.get("actual_amount") is not None:
            actuals_map[(occ["flow_id"], occ["month"])] = occ["actual_amount"]
    
    today = date.today()
    start_of_month = today.replace(day=1)
    end_date = start_of_month + relativedelta(months=horizon)
    expanded = expand_recurring_flows(filtered, start_of_month, end_date)
    
    for ef in expanded:
        fid = ef.get("flow_id", "")
        mkey = ef["date"].strftime("%Y-%m")
        if (fid, mkey) in actuals_map:
            ef["amount"] = actuals_map[(fid, mkey)]
    
    # Group by month, compute net
    months_data = {}
    for i in range(horizon):
        m = start_of_month + relativedelta(months=i)
        key = m.strftime("%Y-%m")
        months_data[key] = {"flows": [], "net": 0.0, "label": m.strftime("%b %Y")}
    
    for ef in expanded:
        mkey = ef["date"].strftime("%Y-%m")
        if mkey in months_data:
            months_data[mkey]["flows"].append(ef)
            months_data[mkey]["net"] += ef["amount"]
    
    # For each negative month, aggregate negative contributors by label
    acc_query = {"entity_id": entity_id} if entity_id else {}
    accounts = await db.bank_accounts.find(acc_query, {"_id": 0}).to_list(100)
    cash_now = sum(acc.get("amount", 0) for acc in accounts)
    
    running = cash_now
    result = []
    for key in sorted(months_data.keys()):
        md = months_data[key]
        running += md["net"]
        
        if md["net"] >= 0:
            continue
        
        # Aggregate only negative flows by label
        label_agg = {}
        for ef in md["flows"]:
            if ef["amount"] >= 0:
                continue
            lbl = ef["label"]
            if lbl not in label_agg:
                label_agg[lbl] = {"label": lbl, "total": 0.0, "count": 0, "category": ef["category"]}
            label_agg[lbl]["total"] += ef["amount"]
            label_agg[lbl]["count"] += 1
        
        # Sort by absolute impact, top 3
        sorted_drivers = sorted(label_agg.values(), key=lambda x: x["total"])[:3]
        drivers = [
            {
                "label": d["label"],
                "amount": round(d["total"], 2),
                "count": d["count"],
                "category": d["category"],
            }
            for d in sorted_drivers
        ]
        
        result.append({
            "month": key,
            "month_label": md["label"],
            "net": round(md["net"], 2),
            "cash_balance": round(running, 2),
            "drivers": drivers,
        })
    
    # Global aggregation: all negative flows across full horizon, aggregated by label
    global_agg = {}
    for key in sorted(months_data.keys()):
        for ef in months_data[key]["flows"]:
            if ef["amount"] >= 0:
                continue
            lbl = ef["label"]
            if lbl not in global_agg:
                global_agg[lbl] = {"label": lbl, "total": 0.0, "count": 0, "category": ef["category"]}
            global_agg[lbl]["total"] += ef["amount"]
            global_agg[lbl]["count"] += 1
    
    global_drivers = [
        {"label": d["label"], "amount": round(d["total"], 2), "count": d["count"], "category": d["category"]}
        for d in sorted(global_agg.values(), key=lambda x: x["total"])[:5]
    ]
    
    return {"negative_months": result, "global_drivers": global_drivers}


@api_router.get("/projection/scenario-delta")
async def get_scenario_delta(
    entity_id: Optional[str] = Query(None),
    horizon: int = Query(12)
):
    """Gap = Likely cash_balance - Committed cash_balance per month. Uses same projection engine."""
    if horizon not in [12, 24, 36]:
        horizon = 12
    
    # Fetch both projections from same engine
    # Committed
    committed = await get_projection(scenario="committed", entity_id=entity_id, horizon=horizon)
    likely = await get_projection(scenario="likely", entity_id=entity_id, horizon=horizon)
    
    months = []
    for cm, lm in zip(committed.months, likely.months):
        gap_net = round(lm.net - cm.net, 2)
        gap_balance = round(lm.closing_cash - cm.closing_cash, 2)
        months.append({
            "month": cm.month,
            "month_label": cm.month_label,
            "committed_net": cm.net,
            "likely_net": lm.net,
            "gap_net": gap_net,
            "committed_balance": cm.closing_cash,
            "likely_balance": lm.closing_cash,
            "gap_balance": gap_balance,
        })
    
    total_gap = round(sum(m["gap_net"] for m in months), 2)
    
    return {
        "months": months,
        "total_gap_net": total_gap,
        "committed_runway": next(
            (i + 1 for i, m in enumerate(committed.months) if m.closing_cash < 0), None
        ),
        "likely_runway": next(
            (i + 1 for i, m in enumerate(likely.months) if m.closing_cash < 0), None
        ),
    }


@api_router.get("/projection/runway")
async def get_cash_runway(
    entity_id: Optional[str] = Query(None),
    horizon: int = Query(36)
):
    """Cash runway = first month where cash_balance < 0. Returns months from now + breach month."""
    if horizon not in [12, 24, 36]:
        horizon = 36
    
    result = {}
    for sc in ["committed", "likely"]:
        proj = await get_projection(scenario=sc, entity_id=entity_id, horizon=horizon)
        breach_idx = None
        breach_month = None
        min_balance = proj.cash_now
        for i, m in enumerate(proj.months):
            if m.closing_cash < min_balance:
                min_balance = m.closing_cash
            if m.closing_cash < 0 and breach_idx is None:
                breach_idx = i + 1
                breach_month = m.month_label
        
        # required_injection = max(0, -min_cash_balance + buffer)
        # buffer defaults to 0, uses system safety_buffer if available
        required_injection = max(0, round(-min_balance, 2))
        
        result[sc] = {
            "months_until_breach": breach_idx,
            "breach_month": breach_month,
            "runway_months": breach_idx if breach_idx else horizon,
            "is_safe": breach_idx is None,
            "min_cash_balance": round(min_balance, 2),
            "required_injection": required_injection,
        }
    
    return result


@api_router.get("/month-details/{month}")
async def get_month_details(
    month: str,
    scenario: str = "likely",
    entity_id: Optional[str] = None
):
    flow_query = {"entity_id": entity_id} if entity_id else {}
    flows = await db.cash_flows.find(flow_query, {"_id": 0}).to_list(1000)
    
    # Dynamic amounts
    flows_with_amounts = []
    for f in flows:
        flows_with_amounts.append(await get_flow_with_dynamic_amount(f))
    
    certainty_levels = get_certainty_levels(scenario)
    filtered = [f for f in flows_with_amounts if f.get("certainty") in certainty_levels]
    
    # Actuals
    all_occs = await db.flow_occurrences.find({}, {"_id": 0}).to_list(1000)
    actuals_map = {}
    occ_detail_map = {}  # (flow_id, month) -> full occurrence record
    for occ in all_occs:
        key = (occ["flow_id"], occ["month"])
        if occ.get("actual_amount") is not None:
            actuals_map[key] = occ["actual_amount"]
        occ_detail_map[key] = occ
    
    year, mon = month.split("-")
    start = date(int(year), int(mon), 1)
    end = start + relativedelta(months=1) - relativedelta(days=1)
    
    # Use the same expand function
    expanded = expand_recurring_flows(filtered, start, end)
    
    # Apply actuals overlay
    for ef in expanded:
        fid = ef.get("flow_id", "")
        mkey = ef["date"].strftime("%Y-%m")
        if (fid, mkey) in actuals_map:
            ef["_planned_amount"] = ef["amount"]
            ef["amount"] = actuals_map[(fid, mkey)]
    
    month_flows = [f for f in expanded if f["date"].strftime("%Y-%m") == month]
    outflows = sorted([f for f in month_flows if f["amount"] < 0], key=lambda x: x["amount"])
    
    recurring_labels = {f["label"] for f in filtered if f.get("recurrence") in ["monthly", "quarterly"]}
    recurring = [f for f in month_flows if f["label"] in recurring_labels]
    
    return {
        "month": month,
        "top_outflows": [{"label": f["label"], "amount": f["amount"], "category": f["category"]} for f in outflows[:5]],
        "recurring_burdens": [{"label": f["label"], "amount": f["amount"], "category": f["category"]} for f in recurring],
        "all_flows": [
            {
                "flow_id": f.get("flow_id", ""),
                "label": f["label"], 
                "amount": round(f["amount"], 2), 
                "planned_amount": round(f.get("_planned_amount", f["amount"]), 2),
                "actual_amount": actuals_map.get((f.get("flow_id", ""), month)),
                "variance_action": occ_detail_map.get((f.get("flow_id", ""), month), {}).get("variance_action"),
                "category": f["category"], 
                "date": f["date"].isoformat(),
                "is_carryover": bool(f.get("carryover_from")),
            } 
            for f in month_flows
        ]
    }

@api_router.get("/variance-summary")
async def get_variance_summary(entity_id: Optional[str] = None):
    """Global variance tracking: total variance, total carried forward, total written off."""
    query = {}
    if entity_id:
        # Get flow IDs for this entity
        entity_flows = await db.cash_flows.find({"entity_id": entity_id}, {"_id": 0, "id": 1}).to_list(1000)
        flow_ids = [f["id"] for f in entity_flows]
        query["flow_id"] = {"$in": flow_ids}
    
    all_occs = await db.flow_occurrences.find(query, {"_id": 0}).to_list(1000)
    
    # Get planned amounts for variance computation
    flow_ids_needed = list({occ["flow_id"] for occ in all_occs if occ.get("actual_amount") is not None})
    flows_map = {}
    if flow_ids_needed:
        flows = await db.cash_flows.find({"id": {"$in": flow_ids_needed}}, {"_id": 0}).to_list(1000)
        for f in flows:
            flows_map[f["id"]] = f
    
    total_variance = 0.0
    total_underperformance = 0.0
    total_overperformance = 0.0
    total_carried = 0.0
    total_written_off = 0.0
    actuals_count = 0
    
    for occ in all_occs:
        if occ.get("actual_amount") is None:
            continue
        actuals_count += 1
        flow = flows_map.get(occ["flow_id"])
        if not flow:
            continue
        planned = flow["amount"]
        rec_mode = flow.get("recurrence_mode", "repeat")
        rec_count = flow.get("recurrence_count")
        if rec_mode == "distribute" and rec_count and rec_count > 0:
            planned = round(planned / rec_count, 2)
        
        variance = planned - occ["actual_amount"]
        abs_var = abs(variance)
        total_variance += abs_var
        
        # Under = actual < planned (for revenue: less came in; for expense: less went out)
        if abs(variance) > 0.01:
            if occ["actual_amount"] < abs(planned):
                total_underperformance += abs_var
            else:
                total_overperformance += abs_var
        
        if occ.get("variance_action") == "carry_forward":
            total_carried += abs_var
        elif occ.get("variance_action") == "write_off":
            total_written_off += abs_var
    
    net_variance_impact = round(total_overperformance - total_underperformance, 2)
    
    return {
        "actuals_recorded": actuals_count,
        "total_variance": round(total_variance, 2),
        "total_underperformance": round(total_underperformance, 2),
        "total_overperformance": round(total_overperformance, 2),
        "total_carried_forward": round(total_carried, 2),
        "total_written_off": round(total_written_off, 2),
        "net_variance_impact": net_variance_impact,
    }


@api_router.get("/")
async def root():
    return {"message": "Cash Piloting Dashboard API"}

# ============== AUTH ENDPOINTS ==============
@api_router.post("/auth/login")
async def login(req: LoginRequest, response: Response):
    email = req.email.strip().lower()
    user = await db.users.find_one({"email": email}, {"_id": 0})
    if not user or not verify_password(req.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    token = create_access_token(user["id"], user["email"])
    response.set_cookie(
        key="access_token", value=token, httponly=True,
        secure=True, samesite="lax", max_age=604800, path="/"
    )
    return {"id": user["id"], "email": user["email"], "name": user.get("name", "")}

@api_router.get("/auth/me")
async def auth_me(user: dict = Depends(get_current_user)):
    return user

@api_router.post("/auth/logout")
async def logout(response: Response):
    response.delete_cookie("access_token", path="/")
    return {"status": "ok"}

# ============== STARTUP ==============
@app.on_event("startup")
async def seed_admin():
    admin_email = os.environ.get("ADMIN_EMAIL", "admin@example.com").strip().lower()
    admin_password = os.environ.get("ADMIN_PASSWORD", "admin123")
    existing = await db.users.find_one({"email": admin_email})
    if existing is None:
        await db.users.insert_one({
            "id": str(uuid.uuid4()),
            "email": admin_email,
            "password_hash": hash_password(admin_password),
            "name": "Admin",
            "created_at": datetime.now(timezone.utc).isoformat()
        })
        logging.info(f"Admin user seeded: {admin_email}")
    elif not verify_password(admin_password, existing["password_hash"]):
        await db.users.update_one(
            {"email": admin_email},
            {"$set": {"password_hash": hash_password(admin_password)}}
        )
        logging.info(f"Admin password updated: {admin_email}")
    await db.users.create_index("email", unique=True)
    await db.actual_import_batches.create_index([("id", 1)], unique=True)
    await db.actual_import_batches.create_index([("entity_id", 1), ("created_at", -1)])
    await db.actual_import_batches.create_index([("status", 1), ("created_at", -1)])
    await db.actual_import_rows.create_index([("id", 1)], unique=True)
    await db.actual_import_rows.create_index([("batch_id", 1), ("row_index", 1)], unique=True)
    await db.actual_import_rows.create_index([("batch_id", 1), ("status", 1)])

app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', 'http://localhost:3000').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
